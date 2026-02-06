"""XLSX to Markdown converter using openpyxl."""

from pathlib import Path

from fileconverter.converters.base import BaseConverter, ConversionResult


class XLSXConverter(BaseConverter):
    """Convert XLSX files to Markdown tables using openpyxl."""

    @classmethod
    def supported_extensions(cls) -> list[str]:
        return [".xlsx"]

    def convert(self, file_path: Path) -> ConversionResult:
        """Convert an XLSX file to Markdown.

        Args:
            file_path: Path to the XLSX file.

        Returns:
            ConversionResult with the conversion outcome.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return self._create_error_result(
                file_path,
                "openpyxl is not installed. Run: pip install openpyxl",
            )

        try:
            workbook = load_workbook(str(file_path), data_only=True)
            markdown_parts: list[str] = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet header
                markdown_parts.append(f"## {sheet_name}\n")

                # Get all rows with data
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                # Find the actual data range (skip completely empty rows/cols)
                rows = self._trim_empty_rows(rows)

                if not rows:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                # Convert to markdown table
                table_md = self._rows_to_markdown_table(rows)
                markdown_parts.append(table_md)
                markdown_parts.append("")

            workbook.close()

            markdown = "\n".join(markdown_parts).strip()
            return self._create_success_result(file_path, markdown)

        except Exception as e:
            return self._create_error_result(file_path, str(e))

    def _trim_empty_rows(self, rows: list[tuple]) -> list[tuple]:
        """Remove completely empty rows from start and end."""
        # Filter out completely empty rows
        non_empty_rows = [
            row for row in rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        if not non_empty_rows:
            return []

        # Find max column width (trim empty columns from right)
        max_col = 0
        for row in non_empty_rows:
            for i, cell in enumerate(row):
                if cell is not None and str(cell).strip():
                    max_col = max(max_col, i + 1)

        # Trim columns
        return [row[:max_col] for row in non_empty_rows]

    def _rows_to_markdown_table(self, rows: list[tuple]) -> str:
        """Convert rows to a markdown table."""
        if not rows:
            return ""

        # Calculate column widths for alignment
        col_count = max(len(row) for row in rows)
        col_widths = [3] * col_count  # Minimum width of 3

        for row in rows:
            for i, cell in enumerate(row):
                if i < col_count:
                    cell_str = self._format_cell(cell)
                    col_widths[i] = max(col_widths[i], len(cell_str))

        lines: list[str] = []

        # Header row (first row)
        header = rows[0]
        header_cells = [
            self._format_cell(header[i] if i < len(header) else "").ljust(col_widths[i])
            for i in range(col_count)
        ]
        lines.append("| " + " | ".join(header_cells) + " |")

        # Separator row
        separator_cells = ["-" * col_widths[i] for i in range(col_count)]
        lines.append("| " + " | ".join(separator_cells) + " |")

        # Data rows
        for row in rows[1:]:
            data_cells = [
                self._format_cell(row[i] if i < len(row) else "").ljust(col_widths[i])
                for i in range(col_count)
            ]
            lines.append("| " + " | ".join(data_cells) + " |")

        return "\n".join(lines)

    def _format_cell(self, cell) -> str:
        """Format a cell value for markdown."""
        if cell is None:
            return ""

        # Convert to string and clean up
        cell_str = str(cell).strip()

        # Escape pipe characters
        cell_str = cell_str.replace("|", "\\|")

        # Replace newlines with spaces
        cell_str = cell_str.replace("\n", " ").replace("\r", "")

        return cell_str
